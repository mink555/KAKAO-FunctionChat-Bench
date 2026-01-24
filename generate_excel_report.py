#!/usr/bin/env python3
"""
FunctionChat-Bench 평가 결과 엑셀 리포트 생성기
- 가독성 최우선
- 간결하고 명확한 구조
"""

import os
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO_PATH = Path(__file__).parent.absolute()
SCORE_PATH = REPO_PATH / "score"
REPORTS_PATH = REPO_PATH / "reports"
REPORTS_PATH.mkdir(exist_ok=True)

# =============================================================================
# 색상 팔레트 (부드럽고 깔끔하게)
# =============================================================================
COLORS = {
    "primary": "2C5282",       # 진한 파랑
    "secondary": "4A90A4",     # 중간 파랑
    "header_bg": "EDF2F7",     # 연한 회색
    "section_bg": "E2E8F0",    # 섹션 배경
    "alt_row": "F7FAFC",       # 교차 행
    "pass_bg": "C6F6D5",       # 연한 초록
    "pass_text": "22543D",     # 진한 초록
    "fail_bg": "FED7D7",       # 연한 빨강
    "fail_text": "742A2A",     # 진한 빨강
    "border": "CBD5E0",        # 테두리
    "text": "2D3748",          # 기본 텍스트
    "text_light": "718096",    # 연한 텍스트
}

# 폰트 (크기 여유있게)
FONTS = {
    "title": Font(bold=True, size=20, color=COLORS["primary"]),
    "subtitle": Font(bold=True, size=14, color=COLORS["secondary"]),
    "section": Font(bold=True, size=13, color=COLORS["primary"]),
    "header": Font(bold=True, size=12, color=COLORS["text"]),
    "normal": Font(size=12, color=COLORS["text"]),
    "small": Font(size=11, color=COLORS["text_light"]),
    "pass": Font(bold=True, size=12, color=COLORS["pass_text"]),
    "fail": Font(bold=True, size=12, color=COLORS["fail_text"]),
}

FILLS = {
    "header": PatternFill(start_color=COLORS["header_bg"], fill_type="solid"),
    "section": PatternFill(start_color=COLORS["section_bg"], fill_type="solid"),
    "alt_row": PatternFill(start_color=COLORS["alt_row"], fill_type="solid"),
    "pass": PatternFill(start_color=COLORS["pass_bg"], fill_type="solid"),
    "fail": PatternFill(start_color=COLORS["fail_bg"], fill_type="solid"),
}

BORDER = Border(
    left=Side(style='thin', color=COLORS["border"]),
    right=Side(style='thin', color=COLORS["border"]),
    top=Side(style='thin', color=COLORS["border"]),
    bottom=Side(style='thin', color=COLORS["border"])
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

# =============================================================================
# 유틸리티
# =============================================================================
def has_evaluation_data(model_dir):
    return len(list(model_dir.glob("*.eval_report.tsv"))) > 0

def set_cell(ws, row, col, value, font=None, fill=None, border=None, align=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if align: cell.alignment = align
    if fmt: cell.number_format = fmt
    return cell

def set_formula(ws, row, col, formula, font=None, fill=None, border=None, align=None, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.value = formula
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if align: cell.alignment = align
    if fmt: cell.number_format = fmt
    return cell

def extract_query(text, max_len=500):
    if pd.isna(text) or not text:
        return ""
    text = str(text)
    try:
        msgs = json.loads(text)
        if isinstance(msgs, list):
            parts = []
            for m in msgs:
                if isinstance(m, dict):
                    role = m.get('role', '')
                    content = m.get('content', '')
                    if role == 'user' and content:
                        parts.append(f"[U] {content}")
                    elif role == 'assistant' and content:
                        parts.append(f"[A] {content}")
            return "\n".join(parts)[:max_len]
    except:
        pass
    return text[:max_len]

def extract_content(text, max_len=300):
    if pd.isna(text) or not text:
        return ""
    text = str(text)
    try:
        data = json.loads(text)
        if isinstance(data, dict):
            if 'tool_calls' in data and data['tool_calls']:
                tc = data['tool_calls'][0]
                func = tc.get('function', {})
                return f"{func.get('name', '')}({func.get('arguments', '')})"
            if 'content' in data:
                return data['content'][:max_len]
    except:
        pass
    return text[:max_len]

def classify_error(reasoning, is_pass):
    if pd.isna(reasoning) or is_pass == "PASS":
        return ""
    text = str(reasoning).lower()
    if 'selection' in text: return "Selection"
    if 'name' in text and 'func' in text: return "Name"
    if 'arg' in text and 'key' in text: return "Arg Key"
    if 'arg' in text and 'value' in text: return "Arg Value"
    if 'halluc' in text: return "Hallucination"
    if 'slot' in text or 'missing' in text: return "Missing Info"
    if 'relevance' in text or 'unnecessary' in text: return "Unnecessary"
    if 'tool_calls' in text and 'null' in text: return "No Call"
    return "Other"

# =============================================================================
# 데이터 수집
# =============================================================================
def collect_model_data(model_name):
    model_name_clean = model_name.replace("/", "_").replace("-", "_")
    org_name = model_name.split('/')[0]
    model_short = model_name.split('/')[-1]
    model_dir = SCORE_PATH / org_name / model_short
    
    if not model_dir.exists() or not has_evaluation_data(model_dir):
        return None
    
    data = {
        "model_name": model_name,
        "model_name_clean": model_name_clean,
        "model_short": model_short,
        "dialog": {},
        "singlecall": {},
        "calldecision": {},
        "all_results": [],
        "error_summary": Counter()
    }
    
    total_pass = 0
    total_count = 0
    
    # Dialog
    dialog_file = model_dir / f"FunctionChat-Dialog.{model_name_clean}.eval_report.tsv"
    if dialog_file.exists():
        try:
            df = pd.read_csv(dialog_file, sep='\t')
            df['is_pass_bool'] = df['is_pass'].astype(str).str.upper() == 'PASS'
            
            for ot in ['call', 'completion', 'slot', 'relevance']:
                type_df = df[df['type_of_output'].astype(str).str.lower() == ot]
                if not type_df.empty:
                    p = int(type_df['is_pass_bool'].sum())
                    t = len(type_df)
                    data["dialog"][ot] = {"pass": p, "total": t, "accuracy": p/t if t > 0 else 0}
                    total_pass += p
                    total_count += t
                    
                    for _, row in type_df.iterrows():
                        is_pass = str(row['is_pass']).upper()
                        err = classify_error(row.get('reasoning', ''), is_pass)
                        data["all_results"].append({
                            "category": f"Dialog-{ot.capitalize()}",
                            "is_pass": is_pass,
                            "id": row.get('#serial_num', ''),
                            "query": extract_query(row.get('query', '')),
                            "gt": extract_content(row.get('ground_truth', '')),
                            "output": extract_content(row.get('model_output', '')),
                            "error": err
                        })
                        if is_pass == "FAIL" and err:
                            data["error_summary"][err] += 1
        except Exception as e:
            print(f"    [WARN] Dialog: {e}")
    
    # SingleCall
    sc_file = model_dir / f"FunctionChat-Singlecall.{model_name_clean}.eval_report.tsv"
    if sc_file.exists():
        try:
            df = pd.read_csv(sc_file, sep='\t')
            df['is_pass_bool'] = df['is_pass'].astype(str).str.upper() == 'PASS'
            
            if 'tools_type' in df.columns:
                for tt in df['tools_type'].unique():
                    type_df = df[df['tools_type'] == tt]
                    p = int(type_df['is_pass_bool'].sum())
                    t = len(type_df)
                    data["singlecall"][tt] = {"pass": p, "total": t, "accuracy": p/t if t > 0 else 0}
                    total_pass += p
                    total_count += t
                    
                    for _, row in type_df.iterrows():
                        is_pass = str(row['is_pass']).upper()
                        err = classify_error(row.get('reasoning', ''), is_pass)
                        data["all_results"].append({
                            "category": f"SingleCall-{tt}",
                            "is_pass": is_pass,
                            "id": row.get('#serial_num', ''),
                            "query": extract_query(row.get('query', '')),
                            "gt": extract_content(row.get('ground_truth', '')),
                            "output": extract_content(row.get('model_output', '')),
                            "error": err
                        })
                        if is_pass == "FAIL" and err:
                            data["error_summary"][err] += 1
        except Exception as e:
            print(f"    [WARN] SingleCall: {e}")
    
    # CallDecision
    cd_files = [
        model_dir / f"FunctionChat-CallDecision-sample.{model_name_clean}.eval_report.tsv",
        model_dir / f"FunctionChat-CallDecision.{model_name_clean}.eval_report.tsv"
    ]
    for cd_file in cd_files:
        if cd_file.exists():
            try:
                df = pd.read_csv(cd_file, sep='\t')
                df['is_pass_bool'] = df['is_pass'].astype(str).str.upper() == 'PASS'
                p = int(df['is_pass_bool'].sum())
                t = len(df)
                data["calldecision"] = {"pass": p, "total": t, "accuracy": p/t if t > 0 else 0}
                total_pass += p
                total_count += t
                
                for _, row in df.iterrows():
                    is_pass = str(row['is_pass']).upper()
                    err = classify_error(row.get('reasoning', ''), is_pass)
                    data["all_results"].append({
                        "category": "CallDecision",
                        "is_pass": is_pass,
                        "id": row.get('#serial_num', ''),
                        "query": extract_query(row.get('query', '')),
                        "gt": extract_content(row.get('ground_truth', '')),
                        "output": extract_content(row.get('model_output', '')),
                        "error": err
                    })
                    if is_pass == "FAIL" and err:
                        data["error_summary"][err] += 1
                break
            except Exception as e:
                print(f"    [WARN] CallDecision: {e}")
    
    data["overall_accuracy"] = total_pass / total_count if total_count > 0 else 0
    data["total_pass"] = total_pass
    data["total_count"] = total_count
    data["total_fail"] = total_count - total_pass
    
    return data

# =============================================================================
# 개별 모델 - Summary 시트
# =============================================================================
def create_model_summary(wb, data):
    ws = wb.active
    ws.title = "Summary"
    r = 1
    
    # 타이틀
    ws.merge_cells(f'A{r}:E{r}')
    set_cell(ws, r, 1, "FunctionChat-Bench Report", font=FONTS["title"])
    ws.row_dimensions[r].height = 35
    r += 2
    
    # 모델 정보
    set_cell(ws, r, 1, "Model", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    ws.merge_cells(f'B{r}:E{r}')
    set_cell(ws, r, 2, data['model_name'], font=FONTS["normal"], border=BORDER, align=ALIGN_LEFT)
    ws.row_dimensions[r].height = 28
    r += 1
    
    set_cell(ws, r, 1, "Date", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    ws.merge_cells(f'B{r}:E{r}')
    set_cell(ws, r, 2, datetime.now().strftime('%Y-%m-%d %H:%M'), font=FONTS["normal"], border=BORDER, align=ALIGN_LEFT)
    ws.row_dimensions[r].height = 28
    r += 2
    
    # Overall (Pass, Fail, Total, Accuracy 순서)
    ws.merge_cells(f'A{r}:E{r}')
    set_cell(ws, r, 1, "Overall Performance", font=FONTS["section"], fill=FILLS["section"])
    ws.row_dimensions[r].height = 30
    r += 1
    
    for c, h in enumerate(["Metric", "Value"], 1):
        set_cell(ws, r, c, h, font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    ws.row_dimensions[r].height = 28
    r += 1
    
    # Pass
    set_cell(ws, r, 1, "Pass", font=FONTS["pass"], border=BORDER, align=ALIGN_LEFT)
    set_formula(ws, r, 2, '=COUNTIF(Details!A:A,"PASS")', font=FONTS["pass"], border=BORDER, align=ALIGN_CENTER)
    ws.row_dimensions[r].height = 26
    pass_row = r
    r += 1
    
    # Fail
    set_cell(ws, r, 1, "Fail", font=FONTS["fail"], border=BORDER, align=ALIGN_LEFT)
    set_formula(ws, r, 2, '=COUNTIF(Details!A:A,"FAIL")', font=FONTS["fail"], border=BORDER, align=ALIGN_CENTER)
    ws.row_dimensions[r].height = 26
    r += 1
    
    # Total
    set_cell(ws, r, 1, "Total", font=FONTS["normal"], border=BORDER, align=ALIGN_LEFT)
    set_formula(ws, r, 2, "=COUNTA(Details!A:A)-1", font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
    ws.row_dimensions[r].height = 26
    total_row = r
    r += 1
    
    # Accuracy
    set_cell(ws, r, 1, "Accuracy", font=FONTS["header"], fill=FILLS["section"], border=BORDER, align=ALIGN_LEFT)
    set_formula(ws, r, 2, f"=IF(B{total_row}>0,B{pass_row}/B{total_row},0)", 
               font=FONTS["header"], fill=FILLS["section"], border=BORDER, align=ALIGN_CENTER, fmt='0.0%')
    ws.row_dimensions[r].height = 28
    r += 2
    
    # Category Performance
    ws.merge_cells(f'A{r}:E{r}')
    set_cell(ws, r, 1, "Category Performance", font=FONTS["section"], fill=FILLS["section"])
    ws.row_dimensions[r].height = 30
    r += 1
    
    headers = ["Group", "Category", "Pass", "Fail", "Total", "Accuracy"]
    for c, h in enumerate(headers, 1):
        set_cell(ws, r, c, h, font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    ws.row_dimensions[r].height = 28
    r += 1
    
    categories = [
        ("Dialog", "Call", "함수 선택 및 인자 추출"),
        ("Dialog", "Completion", "도구 결과를 자연어로 전달"),
        ("Dialog", "Slot", "필수 정보 질문"),
        ("Dialog", "Relevance", "불가 요청에 적절히 응답"),
        ("SingleCall", "exact", "타겟 함수만 제공"),
        ("SingleCall", "4_random", "타겟 + 무작위 3개"),
        ("SingleCall", "4_close", "타겟 + 유사 3개"),
        ("SingleCall", "8_random", "타겟 + 무작위 7개"),
        ("SingleCall", "8_close", "타겟 + 유사 7개"),
        ("Decision", "CallDecision", "호출 여부 판단"),
    ]
    
    cat_start = r
    for group, cat, desc in categories:
        cat_name = f"{group}-{cat}" if group != "Decision" else cat
        
        set_cell(ws, r, 1, group, font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
        set_cell(ws, r, 2, cat, font=FONTS["normal"], border=BORDER, align=ALIGN_LEFT)
        set_formula(ws, r, 3, f'=COUNTIFS(Details!B:B,"{cat_name}",Details!A:A,"PASS")',
                   font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
        set_formula(ws, r, 4, f'=COUNTIFS(Details!B:B,"{cat_name}",Details!A:A,"FAIL")',
                   font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
        set_formula(ws, r, 5, f'=COUNTIF(Details!B:B,"{cat_name}")',
                   font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
        set_formula(ws, r, 6, f"=IF(E{r}>0,C{r}/E{r},0)",
                   font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER, fmt='0.0%')
        
        if r % 2 == 0:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = FILLS["alt_row"]
        ws.row_dimensions[r].height = 26
        r += 1
    
    cat_end = r - 1
    
    # Total row
    set_cell(ws, r, 1, "", font=FONTS["header"], fill=FILLS["header"], border=BORDER)
    set_cell(ws, r, 2, "TOTAL", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_LEFT)
    set_formula(ws, r, 3, f"=SUM(C{cat_start}:C{cat_end})", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    set_formula(ws, r, 4, f"=SUM(D{cat_start}:D{cat_end})", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    set_formula(ws, r, 5, f"=SUM(E{cat_start}:E{cat_end})", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    set_formula(ws, r, 6, f"=IF(E{r}>0,C{r}/E{r},0)", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER, fmt='0.0%')
    ws.row_dimensions[r].height = 28
    r += 2
    
    # Error Analysis
    ws.merge_cells(f'A{r}:E{r}')
    set_cell(ws, r, 1, "Error Analysis", font=FONTS["section"], fill=FILLS["section"])
    ws.row_dimensions[r].height = 30
    r += 1
    
    for c, h in enumerate(["Error Type", "Count", "Ratio"], 1):
        set_cell(ws, r, c, h, font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    ws.row_dimensions[r].height = 28
    r += 1
    
    fail_formula = '=COUNTIF(Details!A:A,"FAIL")'
    
    errors = ["Selection", "Name", "Arg Key", "Arg Value", "Hallucination", "Missing Info", "Unnecessary", "No Call", "Other"]
    for err in errors:
        set_cell(ws, r, 1, err, font=FONTS["normal"], border=BORDER, align=ALIGN_LEFT)
        set_formula(ws, r, 2, f'=COUNTIF(Details!G:G,"{err}")', font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
        set_formula(ws, r, 3, f'=IF({fail_formula}>0,B{r}/{fail_formula},0)', font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER, fmt='0.0%')
        ws.row_dimensions[r].height = 26
        r += 1
    
    # 열 너비
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    
    ws.freeze_panes = "A1"

# =============================================================================
# 개별 모델 - Details 시트
# =============================================================================
def create_model_details(wb, data):
    ws = wb.create_sheet(title="Details")
    r = 1
    
    headers = ["Result", "Category", "ID", "Query", "GT", "Output", "Error"]
    for c, h in enumerate(headers, 1):
        set_cell(ws, r, c, h, font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    ws.row_dimensions[r].height = 30
    r += 1
    
    for item in data["all_results"]:
        if item["is_pass"] == "PASS":
            set_cell(ws, r, 1, "PASS", font=FONTS["pass"], fill=FILLS["pass"], border=BORDER, align=ALIGN_CENTER)
        else:
            set_cell(ws, r, 1, "FAIL", font=FONTS["fail"], fill=FILLS["fail"], border=BORDER, align=ALIGN_CENTER)
        
        set_cell(ws, r, 2, item["category"], font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
        set_cell(ws, r, 3, item["id"], font=FONTS["small"], border=BORDER, align=ALIGN_CENTER)
        set_cell(ws, r, 4, item["query"], font=FONTS["small"], border=BORDER, align=ALIGN_WRAP)
        set_cell(ws, r, 5, item["gt"], font=FONTS["small"], border=BORDER, align=ALIGN_WRAP)
        set_cell(ws, r, 6, item["output"], font=FONTS["small"], border=BORDER, align=ALIGN_WRAP)
        set_cell(ws, r, 7, item["error"], font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
        
        ws.row_dimensions[r].height = 50
        r += 1
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 14
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{r-1}"

# =============================================================================
# 개별 모델 리포트
# =============================================================================
def generate_model_report(model_name):
    data = collect_model_data(model_name)
    if not data:
        return None
    
    model_dir = REPORTS_PATH / data["model_name_clean"]
    model_dir.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    create_model_summary(wb, data)
    create_model_details(wb, data)
    
    output_file = model_dir / f"{data['model_name_clean']}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_file)
    print(f"    [OK] {output_file.name}")
    
    return data

# =============================================================================
# 전체 취합 - Ranking 시트
# =============================================================================
def create_ranking_sheet(wb, all_data):
    ws = wb.active
    ws.title = "Ranking"
    r = 1
    
    ws.merge_cells(f'A{r}:F{r}')
    set_cell(ws, r, 1, "FunctionChat-Bench - Model Ranking", font=FONTS["title"])
    ws.row_dimensions[r].height = 40
    r += 2
    
    ws.merge_cells(f'A{r}:F{r}')
    set_cell(ws, r, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", font=FONTS["small"])
    ws.row_dimensions[r].height = 24
    r += 2
    
    # Pass, Fail, Total, Accuracy 순서
    headers = ["Rank", "Model", "Pass", "Fail", "Total", "Accuracy"]
    for c, h in enumerate(headers, 1):
        set_cell(ws, r, c, h, font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    ws.row_dimensions[r].height = 30
    r += 1
    data_start = r
    
    sorted_data = sorted(all_data, key=lambda x: x['overall_accuracy'], reverse=True)
    
    for rank, m in enumerate(sorted_data, 1):
        short = m['model_short']
        
        set_cell(ws, r, 1, rank, font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
        set_cell(ws, r, 2, m['model_name'], font=FONTS["normal"], border=BORDER, align=ALIGN_LEFT)
        set_formula(ws, r, 3, f'=COUNTIFS(\'All Details\'!B:B,"{short}",\'All Details\'!A:A,"PASS")',
                   font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
        set_formula(ws, r, 4, f'=COUNTIFS(\'All Details\'!B:B,"{short}",\'All Details\'!A:A,"FAIL")',
                   font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
        set_formula(ws, r, 5, f'=COUNTIF(\'All Details\'!B:B,"{short}")',
                   font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
        set_formula(ws, r, 6, f"=IF(E{r}>0,C{r}/E{r},0)", font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER, fmt='0.0%')
        
        if rank == 1:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = FILLS["pass"]
        
        ws.row_dimensions[r].height = 28
        r += 1
    
    data_end = r - 1
    
    # TOTAL
    set_cell(ws, r, 1, "", font=FONTS["header"], fill=FILLS["header"], border=BORDER)
    set_cell(ws, r, 2, "TOTAL", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_LEFT)
    set_formula(ws, r, 3, f"=SUM(C{data_start}:C{data_end})", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    set_formula(ws, r, 4, f"=SUM(D{data_start}:D{data_end})", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    set_formula(ws, r, 5, f"=SUM(E{data_start}:E{data_end})", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    set_formula(ws, r, 6, f"=IF(E{r}>0,C{r}/E{r},0)", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER, fmt='0.0%')
    ws.row_dimensions[r].height = 30
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    
    ws.freeze_panes = "A6"

# =============================================================================
# 전체 취합 - Category Matrix 시트
# =============================================================================
def create_category_matrix(wb, all_data):
    ws = wb.create_sheet(title="Category Matrix")
    r = 1
    
    ws.merge_cells(f'A{r}:H{r}')
    set_cell(ws, r, 1, "Category Performance Matrix", font=FONTS["title"])
    ws.row_dimensions[r].height = 40
    r += 2
    
    categories = [
        ("Dialog", "Call"), ("Dialog", "Completion"), ("Dialog", "Slot"), ("Dialog", "Relevance"),
        ("SingleCall", "exact"), ("SingleCall", "4_random"), ("SingleCall", "4_close"),
        ("SingleCall", "8_random"), ("SingleCall", "8_close"), ("Decision", "CallDecision")
    ]
    
    # 헤더
    set_cell(ws, r, 1, "Group", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    set_cell(ws, r, 2, "Category", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    
    model_shorts = [m['model_short'] for m in all_data]
    for idx, short in enumerate(model_shorts, 3):
        set_cell(ws, r, idx, short[:18], font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    ws.row_dimensions[r].height = 30
    r += 1
    
    for group, cat in categories:
        cat_name = f"{group}-{cat}" if group != "Decision" else cat
        
        set_cell(ws, r, 1, group, font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
        set_cell(ws, r, 2, cat, font=FONTS["normal"], border=BORDER, align=ALIGN_LEFT)
        
        for idx, short in enumerate(model_shorts, 3):
            pass_f = f'COUNTIFS(\'All Details\'!B:B,"{short}",\'All Details\'!C:C,"{cat_name}",\'All Details\'!A:A,"PASS")'
            total_f = f'COUNTIFS(\'All Details\'!B:B,"{short}",\'All Details\'!C:C,"{cat_name}")'
            set_formula(ws, r, idx, f"=IF({total_f}>0,{pass_f}/{total_f},0)",
                       font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER, fmt='0.0%')
        
        if r % 2 == 0:
            for c in range(1, 3 + len(model_shorts)):
                ws.cell(row=r, column=c).fill = FILLS["alt_row"]
        ws.row_dimensions[r].height = 26
        r += 1
    
    # OVERALL
    set_cell(ws, r, 1, "", font=FONTS["header"], fill=FILLS["section"], border=BORDER)
    set_cell(ws, r, 2, "OVERALL", font=FONTS["header"], fill=FILLS["section"], border=BORDER, align=ALIGN_LEFT)
    for idx, short in enumerate(model_shorts, 3):
        pass_f = f'COUNTIFS(\'All Details\'!B:B,"{short}",\'All Details\'!A:A,"PASS")'
        total_f = f'COUNTIF(\'All Details\'!B:B,"{short}")'
        set_formula(ws, r, idx, f"=IF({total_f}>0,{pass_f}/{total_f},0)",
                   font=FONTS["header"], fill=FILLS["section"], border=BORDER, align=ALIGN_CENTER, fmt='0.0%')
    ws.row_dimensions[r].height = 30
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    for idx in range(3, 3 + len(all_data)):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    
    ws.freeze_panes = "C4"

# =============================================================================
# 전체 취합 - Error Summary 시트
# =============================================================================
def create_error_summary(wb, all_data):
    ws = wb.create_sheet(title="Error Summary")
    r = 1
    
    errors = ["Selection", "Name", "Arg Key", "Arg Value", "Hallucination", "Missing Info", "Unnecessary", "No Call", "Other"]
    model_shorts = [m['model_short'] for m in all_data]
    categories = [
        "Dialog-Call", "Dialog-Completion", "Dialog-Slot", "Dialog-Relevance",
        "SingleCall-exact", "SingleCall-4_random", "SingleCall-4_close", 
        "SingleCall-8_random", "SingleCall-8_close", "CallDecision"
    ]
    
    ws.merge_cells(f'A{r}:H{r}')
    set_cell(ws, r, 1, "Error Summary (Category x Model)", font=FONTS["title"])
    ws.row_dimensions[r].height = 36
    r += 2
    
    # 헤더: Category | Error Type | Model1 | Model2 | ... | Total
    total_col = 3 + len(model_shorts)
    set_cell(ws, r, 1, "Category", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    set_cell(ws, r, 2, "Error Type", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    for idx, short in enumerate(model_shorts, 3):
        set_cell(ws, r, idx, short[:16], font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    set_cell(ws, r, total_col, "Total", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    ws.row_dimensions[r].height = 28
    r += 1
    
    data_start = r
    row_idx = 0
    
    for cat in categories:
        cat_start = r
        for err in errors:
            set_cell(ws, r, 1, cat if err == errors[0] else "", font=FONTS["normal"], border=BORDER, align=ALIGN_LEFT)
            set_cell(ws, r, 2, err, font=FONTS["normal"], border=BORDER, align=ALIGN_LEFT)
            
            for idx, short in enumerate(model_shorts, 3):
                set_formula(ws, r, idx, 
                    f'=COUNTIFS(\'All Details\'!B:B,"{short}",\'All Details\'!C:C,"{cat}",\'All Details\'!H:H,"{err}")',
                    font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
            
            last_model_col = get_column_letter(total_col - 1)
            set_formula(ws, r, total_col, f"=SUM(C{r}:{last_model_col}{r})",
                       font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
            
            if row_idx % 2 == 1:
                for c in range(1, total_col + 1):
                    ws.cell(row=r, column=c).fill = FILLS["alt_row"]
            ws.row_dimensions[r].height = 22
            r += 1
            row_idx += 1
        
        # Category subtotal
        set_cell(ws, r, 1, "", font=FONTS["header"], fill=FILLS["section"], border=BORDER)
        set_cell(ws, r, 2, f"{cat} Total", font=FONTS["header"], fill=FILLS["section"], border=BORDER, align=ALIGN_LEFT)
        for idx in range(3, total_col + 1):
            col = get_column_letter(idx)
            set_formula(ws, r, idx, f"=SUM({col}{cat_start}:{col}{r-1})",
                       font=FONTS["header"], fill=FILLS["section"], border=BORDER, align=ALIGN_CENTER)
        ws.row_dimensions[r].height = 26
        r += 1
        row_idx += 1
    
    data_end = r - 1
    
    # Grand Total
    set_cell(ws, r, 1, "", font=FONTS["header"], fill=FILLS["header"], border=BORDER)
    set_cell(ws, r, 2, "GRAND TOTAL", font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_LEFT)
    for idx, short in enumerate(model_shorts, 3):
        set_formula(ws, r, idx, f'=COUNTIFS(\'All Details\'!B:B,"{short}",\'All Details\'!A:A,"FAIL")',
                   font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    last_model_col = get_column_letter(total_col - 1)
    set_formula(ws, r, total_col, f"=SUM(C{r}:{last_model_col}{r})",
               font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    ws.row_dimensions[r].height = 28
    
    # 열 너비
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    for idx in range(3, total_col + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 14
    
    ws.freeze_panes = "C4"

# =============================================================================
# 전체 취합 - All Details 시트
# =============================================================================
def create_all_details(wb, all_data):
    ws = wb.create_sheet(title="All Details")
    r = 1
    
    headers = ["Result", "Model", "Category", "ID", "Query", "GT", "Output", "Error"]
    for c, h in enumerate(headers, 1):
        set_cell(ws, r, c, h, font=FONTS["header"], fill=FILLS["header"], border=BORDER, align=ALIGN_CENTER)
    ws.row_dimensions[r].height = 30
    r += 1
    
    for m in all_data:
        short = m['model_short']
        for item in m["all_results"]:
            if item["is_pass"] == "PASS":
                set_cell(ws, r, 1, "PASS", font=FONTS["pass"], fill=FILLS["pass"], border=BORDER, align=ALIGN_CENTER)
            else:
                set_cell(ws, r, 1, "FAIL", font=FONTS["fail"], fill=FILLS["fail"], border=BORDER, align=ALIGN_CENTER)
            
            set_cell(ws, r, 2, short, font=FONTS["small"], border=BORDER, align=ALIGN_LEFT)
            set_cell(ws, r, 3, item["category"], font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
            set_cell(ws, r, 4, item["id"], font=FONTS["small"], border=BORDER, align=ALIGN_CENTER)
            set_cell(ws, r, 5, item["query"], font=FONTS["small"], border=BORDER, align=ALIGN_WRAP)
            set_cell(ws, r, 6, item["gt"], font=FONTS["small"], border=BORDER, align=ALIGN_WRAP)
            set_cell(ws, r, 7, item["output"], font=FONTS["small"], border=BORDER, align=ALIGN_WRAP)
            set_cell(ws, r, 8, item["error"], font=FONTS["normal"], border=BORDER, align=ALIGN_CENTER)
            
            ws.row_dimensions[r].height = 45
            r += 1
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 14
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{r-1}"

# =============================================================================
# 전체 취합 리포트
# =============================================================================
def create_summary_report(all_data):
    wb = Workbook()
    create_ranking_sheet(wb, all_data)
    create_category_matrix(wb, all_data)
    create_error_summary(wb, all_data)
    create_all_details(wb, all_data)
    
    summary_dir = REPORTS_PATH / "summary"
    summary_dir.mkdir(exist_ok=True)
    output_file = summary_dir / f"All_Models_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_file)
    print(f"    [OK] {output_file.name}")

# =============================================================================
# 메인
# =============================================================================
def main():
    print("=" * 60)
    print("FunctionChat-Bench Report Generator")
    print("=" * 60)
    
    if not SCORE_PATH.exists():
        print("[ERROR] score/ directory not found")
        return
    
    all_models = []
    for org_dir in SCORE_PATH.iterdir():
        if not org_dir.is_dir() or org_dir.name.startswith('.'):
            continue
        for model_dir in org_dir.iterdir():
            if not model_dir.is_dir() or model_dir.name.startswith('.'):
                continue
            if not has_evaluation_data(model_dir):
                print(f"[SKIP] {org_dir.name}/{model_dir.name}")
                continue
            all_models.append(f"{org_dir.name}/{model_dir.name}")
    
    if not all_models:
        print("[ERROR] No evaluated models found")
        return
    
    print(f"\nFound {len(all_models)} model(s)")
    
    all_data = []
    for model_name in all_models:
        print(f"\n[{model_name}] Generating report...")
        data = generate_model_report(model_name)
        if data:
            all_data.append(data)
    
    if all_data:
        print("\n[Summary] Generating consolidated report...")
        create_summary_report(all_data)
    
    print("\n" + "=" * 60)
    print(f"Complete! ({len(all_data)} model(s))")
    print(f"Output: {REPORTS_PATH}")
    print("=" * 60)

if __name__ == "__main__":
    main()
